from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


SUGGESTION_TEMPLATE_LIBRARY = {
    "language_understanding": {
        "优秀": "你读得很认真，能准确抓住句子和段落的意思，表现很棒。",
        "良好": "你已经能读懂文章大意了，继续把句子和段落的重点抓得更准，会越来越棒。",
        "合格": "你已经能看懂不少内容了，再多留意句子里的小细节，会更稳。",
        "待提升": "你已经有基础了，接下来多练习读句子、找段意和联系上下文，就会进步更快。",
    },
    "information_processing": {
        "优秀": "你找重点又快又准，还能把答案整理得清楚又漂亮，真不错。",
        "良好": "你已经能找到大部分关键信息了，继续练习归纳和概括，会越来越厉害。",
        "合格": "你已经能找到一部分信息了，再多练习筛选和整理，答案会更完整。",
        "待提升": "你已经会找一些信息了，接下来多练习找重点、圈关键词，慢慢就会更熟练。",
    },
    "thinking_development": {
        "优秀": "你很会动脑筋，能把原因和结果想得很清楚，思考力很棒。",
        "良好": "你已经能做简单分析和判断了，继续保持，会越来越有条理。",
        "合格": "你已经能看出一些文章里的关系了，再多练习比较和分析，会更有进步。",
        "待提升": "你已经开始会思考了，接下来多练习看关系、想原因和判断对错，会慢慢变强。",
    },
    "literary_appreciation": {
        "优秀": "你很会感受文章里的画面和感情，读书时总能发现很多美好的地方。",
        "良好": "你已经能体会文章中的人物和感情了，继续多读几篇，会越来越细腻。",
        "合格": "你已经能感受到一些内容了，再多关注人物、语言和感情，体会会更丰富。",
        "待提升": "你已经开始留意文章内容了，接下来多体会文章里的人物、语言和感情，审美能力会慢慢提升。",
    },
}


ABILITY_FOCUS_TIPS = {
    "language_understanding": "重点关注词句理解、段落概括和上下文推断。",
    "information_processing": "重点关注关键信息提取、内容筛选和要点归纳。",
    "thinking_development": "重点关注推理判断、结构分析和观点辨析。",
    "literary_appreciation": "重点关注语言美感、形象意境和情感体验。",
}


TAG_SPLIT_PATTERN = re.compile(r"[;,|/、\n]+")


@dataclass
class ParsedQuestion:
    exam_id: str
    book_title: str
    submit_date: str
    question_no: int
    question_id: str
    tags: list[str]
    is_correct: bool
    question_score: float


def normalize_key(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "_")


def coerce_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "是", "对", "正确"}


def coerce_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid numeric value: {value!r}") from exc


def coerce_int(value: Any, default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid integer value: {value!r}") from exc


def split_tags(value: Any) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    return [item.strip() for item in TAG_SPLIT_PATTERN.split(text) if item.strip()]


def sheet_rows_to_dicts(sheet) -> list[dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_key(cell) for cell in rows[0]]
    items: list[dict[str, Any]] = []
    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        item: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            item[header] = row[index] if index < len(row) else None
        items.append(item)
    return items


def find_sheet(workbook, candidates: Iterable[str]):
    normalized = {name.lower(): name for name in workbook.sheetnames}
    for candidate in candidates:
        if candidate.lower() in normalized:
            return workbook[normalized[candidate.lower()]]
    raise ValueError(f"Missing required sheet. Expected one of: {', '.join(candidates)}")


def load_input_from_excel(path: str | Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True)

    student_sheet = find_sheet(workbook, ["student", "students"])
    ability_sheet = find_sheet(workbook, ["ability_catalog", "abilities"])
    tag_map_sheet = find_sheet(workbook, ["tag_ability_map", "mapping"])
    question_sheet = find_sheet(workbook, ["questions", "question_details", "question_records"])

    student_rows = sheet_rows_to_dicts(student_sheet)
    if not student_rows:
        raise ValueError("Student sheet is empty.")
    student_row = student_rows[0]

    ability_catalog = []
    for row in sheet_rows_to_dicts(ability_sheet):
        ability_id = str(row.get("ability_id", "")).strip()
        ability_name = str(row.get("ability_name", "")).strip()
        if ability_id:
            ability_catalog.append({"ability_id": ability_id, "ability_name": ability_name})

    tag_ability_map = []
    for row in sheet_rows_to_dicts(tag_map_sheet):
        tag_name = str(row.get("tag_name", "")).strip()
        ability_id = str(row.get("ability_id", "")).strip()
        if tag_name and ability_id:
            tag_ability_map.append({"tag_name": tag_name, "ability_id": ability_id})

    questions: list[dict[str, Any]] = []
    for row in sheet_rows_to_dicts(question_sheet):
        tags = split_tags(row.get("tag_name"))
        questions.append(
            {
                "exam_id": str(row.get("exam_id", "")).strip(),
                "book_title": str(row.get("book_title", "")).strip(),
                "submit_date": str(row.get("submit_date", "")).strip(),
                "question_no": coerce_int(row.get("question_no")),
                "question_id": str(row.get("question_id", "")).strip(),
                "tags": tags,
                "is_correct": coerce_bool(row.get("is_correct")),
                "question_score": coerce_float(row.get("question_score"), 1.0),
            }
        )

    if not questions:
        raise ValueError("Questions sheet is empty.")

    return {
        "student": {
            "student_id": str(student_row.get("student_id", "")).strip(),
            "student_name": str(student_row.get("student_name", "")).strip(),
        },
        "ability_catalog": ability_catalog,
        "tag_ability_map": tag_ability_map,
        "questions": questions,
    }


def load_input_from_json(path: str | Path) -> dict[str, Any]:
    with open(path, "r", encoding="utf-8") as handle:
        payload = json.load(handle)

    student = payload.get("student", {})
    ability_catalog = payload.get("ability_catalog", [])
    tag_ability_map = payload.get("tag_ability_map", [])
    assessments = payload.get("assessments", [])

    questions: list[dict[str, Any]] = []
    for assessment in assessments:
        exam_id = str(assessment.get("exam_id", "")).strip()
        book_title = str(assessment.get("book_title", "")).strip()
        submit_date = str(assessment.get("submit_date", "")).strip()
        for question in assessment.get("questions", []):
            questions.append(
                {
                    "exam_id": exam_id,
                    "book_title": book_title,
                    "submit_date": submit_date,
                    "question_no": coerce_int(question.get("question_no")),
                    "question_id": str(question.get("question_id", "")).strip(),
                    "tags": split_tags(question.get("tag_name")),
                    "is_correct": coerce_bool(question.get("is_correct")),
                    "question_score": coerce_float(question.get("question_score"), 1.0),
                }
            )

    if not questions:
        raise ValueError("No questions found in JSON input.")

    return {
        "student": {
            "student_id": str(student.get("student_id", "")).strip(),
            "student_name": str(student.get("student_name", "")).strip(),
        },
        "ability_catalog": ability_catalog,
        "tag_ability_map": tag_ability_map,
        "questions": questions,
    }


def load_input(path: str | Path) -> dict[str, Any]:
    suffix = Path(path).suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return load_input_from_excel(path)
    if suffix == ".json":
        return load_input_from_json(path)
    raise ValueError(f"Unsupported input type: {suffix}")


def score_band(score: int) -> str:
    if score >= 90:
        return "优秀"
    if score >= 75:
        return "良好"
    if score >= 60:
        return "合格"
    return "待提升"


def build_suggestion_cards(abilities: list[dict[str, Any]], ability_order: list[str]) -> list[str]:
    ordered_abilities = sorted(
        abilities,
        key=lambda item: (item["score"], ability_order.index(item["ability_id"])),
    )
    selected: list[dict[str, Any]] = []
    selected_ids: set[str] = set()

    for item in ordered_abilities:
        ability_id = item["ability_id"]
        if ability_id in selected_ids:
            continue
        selected.append(item)
        selected_ids.add(ability_id)
        if len(selected) == 2:
            break

    high_band_candidates = [
        item for item in abilities if score_band(item["score"]) in ("优秀", "良好")
    ]
    third_pool = high_band_candidates or abilities
    third_candidates = sorted(
        third_pool,
        key=lambda item: (-item["score"], ability_order.index(item["ability_id"])),
    )
    for item in third_candidates:
        if item["ability_id"] not in selected_ids:
            selected.append(item)
            selected_ids.add(item["ability_id"])
            break

    if len(selected) < 3:
        for item in sorted(
            abilities,
            key=lambda item: (-item["score"], ability_order.index(item["ability_id"])),
        ):
            if item["ability_id"] not in selected_ids:
                selected.append(item)
                selected_ids.add(item["ability_id"])
            if len(selected) == 3:
                break

    suggestion_cards: list[str] = []
    for item in selected[:3]:
        band = item["band"]
        template = SUGGESTION_TEMPLATE_LIBRARY.get(item["ability_id"], {}).get(band)
        if not template:
            template = item["ability_name"]
        suggestion_cards.append(f"{item['ability_name']}：{template}")

    return suggestion_cards



def build_overall_conclusion(overall_score: int, abilities: list[dict[str, Any]]) -> str:
    band = score_band(overall_score)
    if band == "优秀":
        base = "整体表现优秀，四维发展较为均衡，具备较好的阅读理解、信息加工和思维分析能力。"
    elif band == "良好":
        base = "整体表现良好，基础较稳，部分维度已形成优势。"
    elif band == "合格":
        base = "整体能力处于基础发展阶段，部分维度表现稳定，但维度间差异较明显，需针对性训练。"
    else:
        base = "整体阅读能力还有提升空间，建议优先补齐低分维度，逐步夹实基础能力。"

    scores = [item.get("score", 0) for item in abilities]
    if not scores:
        suffix = ""
    else:
        spread = max(scores) - min(scores)
        if spread <= 10:
            suffix = "四维发展较为均衡。"
        elif spread >= 20:
            suffix = "维度间差异较明显，建议优先聚焦低分维度。"
        else:
            suffix = "部分维度存在波动，可继续巩固弱项。"

    return base + suffix


def compute_graph_result(payload: dict[str, Any]) -> dict[str, Any]:
    student = payload["student"]
    ability_catalog = payload["ability_catalog"]
    tag_map_rows = payload["tag_ability_map"]
    questions = payload["questions"]

    tag_to_ability = {str(row["tag_name"]).strip(): str(row["ability_id"]).strip() for row in tag_map_rows}
    ability_order = [str(row["ability_id"]).strip() for row in ability_catalog]
    ability_names = {str(row["ability_id"]).strip(): str(row.get("ability_name", "")).strip() for row in ability_catalog}

    ability_buckets: dict[str, dict[str, float]] = {
        ability_id: {
            "total_weight": 0.0,
            "correct_weight": 0.0,
            "question_count": 0,
            "correct_count": 0,
        }
        for ability_id in ability_order
    }

    mapped_question_count = 0
    total_weight = 0.0
    correct_weight = 0.0
    unmapped_tags: set[str] = set()
    assessment_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for question in questions:
        exam_id = question["exam_id"]
        assessment_groups[exam_id].append(question)

        tags = [tag for tag in question["tags"] if tag]
        if not tags:
            continue

        mapped_tags = []
        for tag in tags:
            ability_id = tag_to_ability.get(tag)
            if ability_id:
                mapped_tags.append((tag, ability_id))
            else:
                unmapped_tags.add(tag)

        if not mapped_tags:
            continue

        mapped_question_count += 1
        weight_per_tag = question["question_score"] / len(tags)

        for _, ability_id in mapped_tags:
            if ability_id not in ability_buckets:
                ability_buckets[ability_id] = {
                    "total_weight": 0.0,
                    "correct_weight": 0.0,
                    "question_count": 0,
                    "correct_count": 0,
                }
                ability_names.setdefault(ability_id, ability_id)
                ability_order.append(ability_id)

            bucket = ability_buckets[ability_id]
            bucket["total_weight"] += weight_per_tag
            bucket["question_count"] += 1
            total_weight += weight_per_tag
            if question["is_correct"]:
                bucket["correct_weight"] += weight_per_tag
                bucket["correct_count"] += 1
                correct_weight += weight_per_tag

    abilities = []
    for ability_id in ability_order:
        bucket = ability_buckets.get(ability_id, {})
        total = float(bucket.get("total_weight", 0.0))
        correct = float(bucket.get("correct_weight", 0.0))
        score = round((correct / total) * 100) if total > 0 else 0
        accuracy = round((correct / total), 4) if total > 0 else 0.0
        abilities.append(
            {
                "ability_id": ability_id,
                "ability_name": ability_names.get(ability_id, ability_id),
                "score": score,
                "band": score_band(score),
                "question_count": int(bucket.get("question_count", 0)),
                "correct_count": int(bucket.get("correct_count", 0)),
                "accuracy": accuracy,
            }
        )

    overall_score = round((correct_weight / total_weight) * 100) if total_weight > 0 else 0
    overall_band = score_band(overall_score)
    score_spread = max((item["score"] for item in abilities), default=0) - min((item["score"] for item in abilities), default=0)
    radar_values = [item["score"] for item in abilities]

    suggestions = build_suggestion_cards(abilities, ability_order)

    trend = []
    for exam_id, group in sorted(assessment_groups.items(), key=lambda kv: _assessment_sort_key(kv[1][0])):
        exam_total = 0.0
        exam_correct = 0.0
        for question in group:
            tags = [tag for tag in question["tags"] if tag]
            if not tags:
                continue
            mapped_tags = [tag for tag in tags if tag in tag_to_ability]
            if not mapped_tags:
                continue
            weight_per_tag = question["question_score"] / len(tags)
            mapped_weight = weight_per_tag * len(mapped_tags)
            exam_total += mapped_weight
            if question["is_correct"]:
                exam_correct += mapped_weight

        trend.append(
            {
                "period": group[0]["submit_date"],
                "book_title": group[0]["book_title"],
                "score": round((exam_correct / exam_total) * 100) if exam_total > 0 else 0,
            }
        )

    assessment_count = len(trend)
    confidence_level = _confidence_level(mapped_question_count, assessment_count)
    overall_conclusion = build_overall_conclusion(overall_score, abilities)

    result = {
        "student": student,
        "overall_score": overall_score,
        "overall_band": overall_band,
        "overall_conclusion": overall_conclusion,
        "score_spread": score_spread,
        "sample_count": mapped_question_count,
        "assessment_count": assessment_count,
        "confidence_level": confidence_level,
        "abilities": abilities,
        "radar_values": radar_values,
        "suggestions": suggestions,
        "trend": trend,
        "unmapped_tags": sorted(unmapped_tags),
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    return result


def _assessment_sort_key(question: dict[str, Any]) -> tuple[str, int]:
    return (str(question.get("submit_date", "")), coerce_int(question.get("question_no")))


def _confidence_level(sample_count: int, assessment_count: int) -> str:
    if sample_count >= 80 and assessment_count >= 5:
        return "high"
    if sample_count >= 40 and assessment_count >= 3:
        return "medium"
    return "low"


def markdown_table(headers: list[str], rows: list[list[Any]]) -> str:
    header_line = "| " + " | ".join(headers) + " |"
    separator = "| " + " | ".join(["---"] * len(headers)) + " |"
    data_lines = ["| " + " | ".join(str(cell) for cell in row) + " |" for row in rows]
    return "\n".join([header_line, separator, *data_lines])


def generate_report_markdown(result: dict[str, Any]) -> str:
    student = result["student"]
    abilities = result["abilities"]
    trend = result["trend"]
    unmapped_tags = result.get("unmapped_tags", [])

    lines = [
        "# 学生阅读能力图谱报告",
        "",
        f"**学生**?{student.get('student_name', '')}",
        f"**学生ID**?{student.get('student_id', '')}",
        "",
        "## 1. 综合阅读能力分",
        "",
        f"**{result['overall_score']} ?**",
        f"趋势判断：{result.get('overall_band', score_band(result['overall_score']))}",
        "",
        "趋势判断：趋势判断：趋势判断：趋势判断：",
        "",
        f"## 2. {len(abilities)} ?能力维度??",
        "",
    ]

    ability_rows = [
        [
            item["ability_name"],
            item["score"],
            item["band"],
            item["question_count"],
            item["correct_count"],
            f"{item['accuracy'] * 100:.1f}%",
        ]
        for item in abilities
    ]
    lines.append(markdown_table(["????", "??", "??", "??", "???", "???"], ability_rows))

    lines.extend(
        [
            "",
            "## 3. 总体结论",
            "",
            result.get("overall_conclusion", build_overall_conclusion(result["overall_score"], abilities)),
            "",
            "## 4. 雷达图数据",
            "",
            "? `ability_catalog` 趋势判断：",
            "",
            "```json",
            json.dumps(result["radar_values"], ensure_ascii=False),
            "```",
            "",
            "## 5. 薄弱项分析",
            "",
        ]
    )

    weakest = sorted(abilities, key=lambda item: item["score"])[:2]
    for index, item in enumerate(weakest, start=1):
        lines.extend(
            [
                f"{index}. **{item['ability_name']}**",
                f"   - ?? {item['score']} ??",
                f"   - 趋势判断：{item['band']}?",
                "   - 趋势判断：趋势判断：???",
                f"   - {ABILITY_FOCUS_TIPS.get(item['ability_id'], '趋势判断：趋势判断：????')}",
                "",
            ]
        )

    lines.extend(
        [
            "## 6. 提升建议",
            "",
        ]
    )

    for suggestion in result["suggestions"]:
        lines.append(f"- {suggestion}")

    lines.extend(
        [
            "",
            "## 7. 成长趋势",
            "",
            markdown_table(
                ["????", "????", "???"],
                [[item["period"], item["book_title"], item["score"]] for item in trend],
            ),
            "",
            "趋势判断：",
            "- 趋势判断：趋势判断：趋势判断：??",
            "- 趋势判断：趋势判断：趋势判断：?",
            "",
            "## 8. 数据说明",
            "",
            markdown_table(
                ["??", "?"],
                [
                    ["????", result.get("assessment_count", 0)],
                    ["趋势判断：", result.get("sample_count", 0)],
                    ["???", result.get("confidence_level", "")],
                    ["????", result.get("overall_band", "")],
                    ["????", result.get("score_spread", 0)],
                    ["趋势判断：?", result.get("updated_at", "")],
                ],
            ),
        ]
    )

    if unmapped_tags:
        lines.extend(
            [
                "",
                "## 9. 趋势判断：",
                "",
                "- " + "?".join(unmapped_tags),
            ]
        )

    return "\n".join(lines).rstrip() + "\n"


def run(
input_path: str | Path, output_md: str | Path | None = None, output_json: str | Path | None = None) -> dict[str, Any]:
    payload = load_input(input_path)
    result = compute_graph_result(payload)
    markdown = generate_report_markdown(result)

    if output_md:
        Path(output_md).write_text(markdown, encoding="utf-8")
    if output_json:
        Path(output_json).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate reading ability graph report from Excel or JSON input.")
    parser.add_argument("--input", required=True, help="Input Excel or JSON file.")
    parser.add_argument("--output-md", help="Write Markdown report to this file.")
    parser.add_argument("--output-json", help="Write computed result JSON to this file.")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_arg_parser().parse_args(argv)
    result = run(args.input, args.output_md, args.output_json)
    if not args.output_md:
        print(generate_report_markdown(result))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
