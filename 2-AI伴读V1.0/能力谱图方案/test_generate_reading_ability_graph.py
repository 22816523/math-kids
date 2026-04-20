import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from generate_reading_ability_graph import (
    compute_graph_result,
    generate_report_markdown,
    load_input_from_excel,
)


class ReadingAbilityGraphTests(unittest.TestCase):
    def build_balanced_workbook(self, path: Path) -> None:
        wb = Workbook()

        ws = wb.active
        ws.title = "student"
        ws.append(["student_id", "student_name"])
        ws.append(["S001", "小明"])

        ws = wb.create_sheet("ability_catalog")
        ws.append(["ability_id", "ability_name"])
        ws.append(["comprehension", "理解能力"])
        ws.append(["analysis", "分析能力"])
        ws.append(["expression", "表达能力"])
        ws.append(["critical_thinking", "思辨能力"])
        ws.append(["creativity", "创造能力"])
        ws.append(["memory", "记忆能力"])

        ws = wb.create_sheet("tag_ability_map")
        ws.append(["tag_name", "ability_id"])
        ws.append(["事实提取", "comprehension"])
        ws.append(["人物关系", "comprehension"])
        ws.append(["细节记忆", "memory"])
        ws.append(["情节梳理", "analysis"])
        ws.append(["因果推理", "critical_thinking"])
        ws.append(["主旨概括", "expression"])
        ws.append(["观点判断", "critical_thinking"])
        ws.append(["联想迁移", "creativity"])

        ws = wb.create_sheet("questions")
        ws.append(
            [
                "exam_id",
                "book_title",
                "submit_date",
                "question_no",
                "question_id",
                "tag_name",
                "is_correct",
                "question_score",
            ]
        )
        rows = [
            ["X001", "《三体》", "2026-03-01", 1, "X001-01", "事实提取", 1, 5],
            ["X001", "《三体》", "2026-03-01", 2, "X001-02", "情节梳理", 0, 5],
            ["X001", "《三体》", "2026-03-01", 3, "X001-03", "主旨概括", 1, 5],
            ["X001", "《三体》", "2026-03-01", 4, "X001-04", "因果推理", 0, 5],
            ["X001", "《三体》", "2026-03-01", 5, "X001-05", "联想迁移", 1, 5],
            ["X001", "《三体》", "2026-03-01", 6, "X001-06", "细节记忆", 0, 5],
            ["X002", "《西游记》", "2026-03-08", 1, "X002-01", "人物关系", 0, 5],
            ["X002", "《西游记》", "2026-03-08", 2, "X002-02", "情节梳理", 1, 5],
            ["X002", "《西游记》", "2026-03-08", 3, "X002-03", "主旨概括", 0, 5],
            ["X002", "《西游记》", "2026-03-08", 4, "X002-04", "观点判断", 1, 5],
            ["X002", "《西游记》", "2026-03-08", 5, "X002-05", "联想迁移", 0, 5],
            ["X002", "《西游记》", "2026-03-08", 6, "X002-06", "细节记忆", 1, 5],
        ]
        for row in rows:
            ws.append(row)

        wb.save(path)

    def build_multi_tag_workbook(self, path: Path) -> None:
        wb = Workbook()

        ws = wb.active
        ws.title = "student"
        ws.append(["student_id", "student_name"])
        ws.append(["S001", "小明"])

        ws = wb.create_sheet("ability_catalog")
        ws.append(["ability_id", "ability_name"])
        ws.append(["comprehension", "理解能力"])
        ws.append(["memory", "记忆能力"])

        ws = wb.create_sheet("tag_ability_map")
        ws.append(["tag_name", "ability_id"])
        ws.append(["事实提取", "comprehension"])
        ws.append(["细节记忆", "memory"])

        ws = wb.create_sheet("questions")
        ws.append(
            [
                "exam_id",
                "book_title",
                "submit_date",
                "question_no",
                "question_id",
                "tag_name",
                "is_correct",
                "question_score",
            ]
        )
        ws.append(["X001", "《三体》", "2026-03-01", 1, "X001-01", "事实提取;细节记忆", 1, 6])
        ws.append(["X001", "《三体》", "2026-03-01", 2, "X001-02", "不存在的标签", 1, 5])

        wb.save(path)

    def test_compute_graph_result_from_excel(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "input.xlsx"
            self.build_balanced_workbook(workbook_path)

            parsed = load_input_from_excel(workbook_path)
            result = compute_graph_result(parsed)
            markdown = generate_report_markdown(result)

            self.assertEqual(result["student"]["student_name"], "小明")
            self.assertEqual(result["overall_score"], 50)
            self.assertEqual(result["confidence_level"], "low")
            self.assertEqual(result["radar_values"], [50, 50, 50, 50, 50, 50])
            self.assertEqual(result["unmapped_tags"], [])
            self.assertEqual(result["trend"][0]["score"], 50)
            self.assertIn("# 学生阅读能力图谱报告", markdown)
            self.assertIn("综合阅读能力分", markdown)
            self.assertIn("雷达图数据", markdown)
            self.assertIn("薄弱项分析", markdown)

    def test_multi_tag_weight_split(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "input.xlsx"
            self.build_multi_tag_workbook(workbook_path)

            parsed = load_input_from_excel(workbook_path)
            result = compute_graph_result(parsed)

            self.assertIn("不存在的标签", result["unmapped_tags"])
            self.assertEqual(result["abilities"][0]["ability_id"], "comprehension")
            self.assertEqual(result["abilities"][0]["score"], 100)
            self.assertEqual(result["abilities"][1]["ability_id"], "memory")
            self.assertEqual(result["abilities"][1]["score"], 100)


if __name__ == "__main__":
    unittest.main()
